import os, re, io
from datetime import datetime
import xml.etree.ElementTree as ET
import pdfplumber
import pandas as pd
from flask import Flask, render_template, request, redirect, url_for, jsonify, send_file
from flask_sqlalchemy import SQLAlchemy
from openpyxl.styles import Alignment, Border, Side

#flash应用与配置
app = Flask(__name__)
base_dir = os.path.dirname(os.path.abspath(__file__))
app.config['SQLALCHEMY_DATABASE_URI'] = 'sqlite:///' + os.path.join(base_dir, 'reimbursement.db')
app.config['UPLOAD_FOLDER'] = os.path.join(base_dir, 'static/uploads')
db = SQLAlchemy(app)

os.makedirs(app.config['UPLOAD_FOLDER'], exist_ok=True)

# --- 数据库模型 ---
class Invoice(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    inv_date = db.Column(db.String(50))
    seller = db.Column(db.String(200))
    amount = db.Column(db.Float)
    file_path = db.Column(db.String(200))
    category = db.Column(db.String(100))
    content = db.Column(db.String(200))
    claimant = db.Column(db.String(100), default="") # 新增：发票行报销人
    claim_id = db.Column(db.Integer, db.ForeignKey('claim.id'))

class Claim(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    employee_name = db.Column(db.String(50))
    total_amount = db.Column(db.Float)
    claim_date = db.Column(db.String(50), default="") # 新增：报销日期字段
    create_time = db.Column(db.DateTime, default=datetime.now)
    invoices = db.relationship('Invoice', backref='claim', lazy=True, order_by="Invoice.inv_date")

with app.app_context():
    db.create_all()

# --- 功能函数 ---
#从发票中识别金额
def parse_pdf_amount(pdf_path):
    amount = "0.00"
    try:
        with pdfplumber.open(pdf_path) as pdf:
            page = pdf.pages[0]
            text = page.extract_text()
            patterns = [
                r'\(小写\)\s*[￥¥]?\s*([0-9]+\.[0-9]{2})',
                r'（小写）\s*[￥¥]?\s*([0-9]+\.[0-9]{2})',
                r'价税合计\s*[￥¥]?\s*([0-9]+\.[0-9]{2})',
                r'[￥¥]\s*([0-9]+\.[0-9]{2})'
            ]
            for p in patterns:
                matches = re.findall(p, text)
                if matches:
                    amount = matches[-1]
                    break
    except: pass
    return amount
#文件名中识别销售方
def extract_seller(filename):
    chinese = "".join(re.findall(r'[\u4e00-\u9fa5（）\(\)]+', filename))
    return chinese if chinese else "未知销售方"

# --- 路由接口 ---
@app.route('/')
def index():
    unclaimed = Invoice.query.filter_by(claim_id=None).order_by(Invoice.inv_date.asc()).all()
    # 提取科目和报销人用于前端筛选器
    cats = sorted(list(set([i.category for i in unclaimed if i.category])))
    clas = sorted(list(set([i.claimant for i in unclaimed if i.claimant])))
    return render_template('input.html', invoices=unclaimed, categories=cats, claimants=clas)

@app.route('/upload_batch', methods=['POST'])
def upload_batch():
    files = request.files.getlist('files')
    results = []
    for file in files:
        filename = datetime.now().strftime("%Y%m%d%H%M%S_") + file.filename
        path = os.path.join(app.config['UPLOAD_FOLDER'], filename)
        file.save(path)
        info = {"date": "", "amount": "0.00", "seller": extract_seller(file.filename)}
        if not file.filename.lower().endswith('.xml'):
            info["amount"] = parse_pdf_amount(path)
            try:
                with pdfplumber.open(path) as pdf:
                    text = pdf.pages[0].extract_text()
                    d = re.search(r'(\d{4}[年-]\d{2}[月-]\d{2}[日]?)', text)
                    if d: info["date"] = d.group(1).replace('年','-').replace('月','-').replace('日','')
            except: pass
        info.update({'filename': filename, 'original_name': file.filename, 'category': '', 'content': '', 'claimant': ''})
        results.append(info)
    return jsonify(results)

@app.route('/save_invoices', methods=['POST'])
def save_invoices():
    for item in request.json:
        inv = Invoice(inv_date=item['date'], seller=item['seller'], amount=float(item['amount'] or 0),
                      category=item['category'], content=item['content'], claimant=item.get('claimant',''), file_path=item['filename'])
        db.session.add(inv)
    db.session.commit()
    return jsonify({"status": "success"})

@app.route('/update_invoice_detail', methods=['POST'])
def update_invoice_detail():
    data = request.json
    inv = db.session.get(Invoice, data['id'])
    if inv:
        inv.amount = float(data['amount'])
        inv.category = data['category']
        inv.content = data['content']
        inv.claimant = data.get('claimant', '')
        db.session.commit()
        return jsonify({"status": "success"})
    return jsonify({"status": "error"}), 404

@app.route('/delete_invoice/<int:id>')
def delete_invoice(id):
    inv = db.session.get(Invoice, id)
    if inv:
        db.session.delete(inv)
        db.session.commit()
    return redirect(url_for('index'))


# 【新增/修改】删除报销单路由
@app.route('/delete_claim/<int:id>')
def delete_claim(id):
    claim = db.session.get(Claim, id)
    if claim:
        # 核心逻辑：遍历该单据下的所有发票，将它们的 claim_id 置为空
        # 这样它们就会重新出现在“待处理列表”中
        for inv in claim.invoices:
            inv.claim_id = None

        # 然后删除报销单本身
        db.session.delete(claim)
        db.session.commit()
        return redirect(url_for('output'))

    return "未找到该报销单", 404

@app.route('/update_claim_date', methods=['POST'])
def update_claim_date():
    data = request.json
    claim = db.session.get(Claim, data['id'])
    if claim:
        claim.claim_date = data['date']
        db.session.commit()
        return jsonify({"status": "success"})
    return jsonify({"status": "error"}), 404

@app.route('/create_claim', methods=['POST'])
def create_claim():
    ids = request.form.getlist('selected_invoices')
    emp_name = request.form.get('employee_name')
    if not ids or not emp_name: return redirect(url_for('index'))
    new_claim = Claim(employee_name=emp_name, total_amount=0)
    db.session.add(new_claim); db.session.flush()
    total = sum([db.session.get(Invoice, int(i_id)).amount for i_id in ids])
    for i_id in ids:
        inv = db.session.get(Invoice, int(i_id))
        inv.claim_id = new_claim.id
    new_claim.total_amount = round(total, 2)
    db.session.commit()
    return redirect(url_for('output'))


@app.route('/merge_claims', methods=['POST'])
def merge_claims():
    selected_ids = request.form.getlist('selected_claims')
    if not selected_ids or len(selected_ids) < 2:
        return redirect(url_for('output'))

    # 1. 获取所有待处理的报销单对象
    claims_to_merge = Claim.query.filter(Claim.id.in_(selected_ids)).all()

    # 2. 按姓名分组合并
    name_groups = {}
    for c in claims_to_merge:
        if c.employee_name not in name_groups:
            name_groups[c.employee_name] = []
        name_groups[c.employee_name].append(c)

    for name, group in name_groups.items():
        if len(group) < 2: continue  # 只有一个人有多张单据时才合并

        # 指定第一张单据为“主单据”
        main_claim = group[0]
        other_claims = group[1:]
        other_ids = [oc.id for oc in other_claims]

        # 3. 【关键修复】使用批量更新：将所有属于“其他单据”的发票全部改派到“主单据”下
        Invoice.query.filter(Invoice.claim_id.in_(other_ids)).update({Invoice.claim_id: main_claim.id},
                                                                     synchronize_session=False)

        # 4. 重新计算主单据总额
        new_total = sum(oc.total_amount for oc in group)
        main_claim.total_amount = round(new_total, 2)

        # 5. 删除已被合并掉的空单据
        for oc in other_claims:
            db.session.delete(oc)

    db.session.commit()  # 统一提交事务
    return redirect(url_for('output'))

@app.route('/output')
def output():
    claims = Claim.query.order_by(Claim.create_time.desc()).all()
    return render_template('output.html', claims=claims)

@app.route('/export_excel', methods=['POST'])
def export_excel():
    selected_ids = request.form.getlist('selected_claims')
    if not selected_ids: return "请勾选报销单", 400
    output_stream = io.BytesIO()
    with pd.ExcelWriter(output_stream, engine='openpyxl') as writer:
        all_data = []
        summary_rows = []
        current_row = 2
        for c_id in selected_ids:
            claim = db.session.get(Claim, int(c_id))
            if not claim: continue
            for i, inv in enumerate(claim.invoices):
                all_data.append([i + 1, claim.employee_name, inv.inv_date, inv.seller, inv.category, inv.content, inv.amount])
                current_row += 1
            summary_val = f"报销单汇总金额：￥{claim.total_amount}"
            all_data.append(["总金额", summary_val, "", "", "", "", ""])
            summary_rows.append(current_row)
            current_row += 2
            all_data.append(["", "", "", "", "", "", ""])
        df = pd.DataFrame(all_data, columns=['序号', '报销人姓名', '开票日期', '销售方', '科目', '内容', '金额'])
        df.to_excel(writer, index=False, sheet_name='报销汇总')
        ws = writer.sheets['报销汇总']
        for row_num in summary_rows:
            ws.merge_cells(start_row=row_num, start_column=2, end_row=row_num, end_column=7)
            ws.cell(row=row_num, column=2).alignment = Alignment(horizontal='center')
            ws.cell(row=row_num, column=1).alignment = Alignment(horizontal='center')
        column_widths = [8, 15, 15, 30, 15, 30, 12]
        for i, width in enumerate(column_widths, start=1):
            ws.column_dimensions[ws.cell(row=1, column=i).column_letter].width = width
    output_stream.seek(0)
    return send_file(output_stream, as_attachment=True, download_name=f"报销汇总_{datetime.now().strftime('%Y%m%d')}.xlsx")

if __name__ == '__main__':
    app.run(debug=True)